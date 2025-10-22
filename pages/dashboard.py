import dash
from dash import html, dcc, Input, Output, State, callback, ALL, MATCH, ctx
from dash.exceptions import PreventUpdate
import dash_bootstrap_components as dbc
import pandas as pd
from models import Grant, User
from db_utils import get_db_session
from flask_login import current_user
import io
import base64
from datetime import datetime
from sqlalchemy import distinct, asc, func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from models import Grant, GrantPersonnel, GrantTravel,TravelItinerary, GrantMaterial, ExpenseCategory, ExpenseSubcategory, PI, CoPI, ProfessionalStaff, Postdoc, GRA, Undergrad, TempHelp, NSFPersonnelCompensation, NIHPersonnelCompensation, NSFFringeRate, NIHFringeRate, GraduateStudentCost
from dash import dash_table
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.styles import ParagraphStyle



# Register the page
dash.register_page(__name__, path='/dashboard')

# Define the layout
def layout():
    return html.Div([
        dcc.Store(id="selected-grant-id"),
        dcc.Store(id="view-grant-id-store"),
        dcc.Location(id="redirect-location", refresh=True),
        dcc.Download(id="download-pdf"), 
        dcc.Download(id="download-excel-modal"),


        html.H3('Available Budgets', className="text-center mt-4 mb-4", style={"color": "#2c3e50"}),

        dbc.Container([
            dbc.Row([
                dbc.Col([
                    html.Div(id='table-container'),
                ], width=12),
            ], className="mb-4"),


            # Hidden div for storing download data
            html.Div(id='download-data', style={'display': 'none'}),
            dcc.Download(id="download-excel")
        ], fluid=True), 

        # modal to view the grant information
        dbc.Modal(
            [
                dbc.ModalHeader("Grant Details"),
                dbc.ModalBody(id="view-grant-modal-body"),
                dbc.ModalFooter([
                    dbc.Button("Download Details as Excel", id="download-modal-excel-btn", n_clicks=0, color="success", className="me-2"),
                    dbc.Button("Download Details as PDF",   id="download-modal-pdf-btn", n_clicks=0,   color="secondary", className="me-2"),
                    dbc.Button("Close", id="close-view-modal", className="ms-auto", n_clicks=0),
                ]),
            ],
            id="view-grant-modal",
            is_open=False,
            size="xl",  # large modal
        ),
    ])

# Callback to fetch and display grants
@callback(
    Output('table-container', 'children'),
    Input('table-container', 'id'),  # Trigger on button click
    prevent_initial_call=False
)
def display_grants(id):
    # Get the database session
    session = get_db_session()

    try:
        # Ensure the current user is logged in
        if not current_user.is_authenticated:
            return html.Div("You must be logged in to view your grants.", style={"color": "red"})

        # Fetch the grants created by the current user
        grants = session.query(Grant).filter(Grant.user_id == current_user.id).all()

        # If no grants found, return a message with an image
        if not grants:
            return html.Div([
                html.Img(src='assets/images/nogrants1.png', style={"width": "500px", "height":"100%", "display": "block", "margin": "auto"}),
                html.Div("No grants found.", style={"color": "gray", "text-align": "center", "margin-top": "10px"})
            ])

        # Prepare data for displaying in rows and columns
        rows = []
        for idx, grant in enumerate(grants, 1):
            # Determine if the row index is even
            row_style = {'backgroundColor': '#f8f9fa'} if idx % 2 == 0 else {}
            rows.append(
                dbc.Row([
                    # dbc.Col(html.Div(str(idx)), width=1),
                    dbc.Col(html.Div(grant.title), width=2),
                    dbc.Col(html.Div(grant.created_at.strftime('%Y-%m-%d')), width=2),
                    dbc.Col(html.Div(grant.funding_agency), width=1),
                    dbc.Col(html.Div(grant.start_date.strftime('%Y-%m-%d')), width=2),
                    dbc.Col(html.Div(grant.end_date.strftime('%Y-%m-%d')), width=2),
                    dbc.Col(html.Div([
                        dbc.Button("Edit", n_clicks=0,  id={'type': 'edit-btn', 'index': grant.id}, className="btn btn-warning btn-sm me-2"),
                        dbc.Button("View", n_clicks=0,  id={'type': 'manage-btn', 'index': grant.id}, className="btn btn-info btn-sm me-2"),
                        dbc.Button("Download Excel", n_clicks=0,  id={'type': 'download-excel-btn', 'index': grant.id}, className="btn btn-success btn-sm me-2", ), 
                        dbc.Button("Delete", n_clicks=0,  id={'type': 'delete-btn', 'index': grant.id}, className="btn btn-danger btn-sm me-2"),
                    ], style={"display": "flex", "gap": "0px", }), width=3)
                ], className="mb-2 p-2 rounded", style=row_style)
            )

        # Display the table
            table = html.Div([
                dbc.Row([
                    # dbc.Col(html.Div("SN", className="text-center p-2"), width=1),
                    dbc.Col(html.Div("Grant Name", className="text-left p-2"), width=2),
                    dbc.Col(html.Div("Date Created", className="text-left p-2"), width=2),
                    dbc.Col(html.Div("Funding Agency", className="text-left p-2"), width=1),
                    dbc.Col(html.Div("Start Date", className="text-left p-2"), width=2),
                    dbc.Col(html.Div("End Date", className="text-left p-2"), width=2),
                    dbc.Col(html.Div("Actions", className="text-left p-2"), width=3)
                ], className="fw-bold mb-2 bg-light border rounded ", style={"font-size": "1rem", }),  # Styled header row
                *rows  # Adding all the rows dynamically
            ])


        return table

    except Exception as err:
        return html.Div(f"Error: {str(err)}", style={"color": "red"})

    finally:
        session.close()  # Close the session




##############################################
# view the modal
##############################################
@callback(
    Output("view-grant-modal", "is_open"),
    Output("view-grant-modal-body", "children"),
    Output("view-grant-id-store", "data"),
    Input({"type": "manage-btn", "index": ALL}, "n_clicks"),
    Input("close-view-modal", "n_clicks"),
    State("view-grant-modal", "is_open"),
    prevent_initial_call=True
)
def toggle_view_modal(n_clicks_list, close_clicks, is_open):
    triggered = ctx.triggered_id
    if not (triggered and any(n_clicks_list)):
        return is_open, dash.no_update, dash.no_update
    
    # If they clicked “Close”
    if triggered == "close-view-modal":
        # hide modal, leave body untouched
        return False, dash.no_update, dash.no_update

    grant_id = triggered["index"]
    session = get_db_session()
    grant = session.query(Grant).get(grant_id)

    # 1) Grant Info
    pis = [r[0] for r in session
       .query(distinct(GrantPersonnel.name))
       .filter_by(grant_id=grant_id, position="PI")
       .all()]
    copis = [r[0] for r in session
         .query(distinct(GrantPersonnel.name))
         .filter_by(grant_id=grant_id, position="Co-PI")
         .all()]
    grant_info = html.Div([
        html.H5(grant.title),
        html.P([html.Strong("Funding Agency: "), grant.funding_agency]),
        html.P([html.Strong("Period: "), f"{grant.start_date} → {grant.end_date}"]),
        html.P([html.Strong("Duration: "), f"{grant.duration} years"]),
        html.P([html.Strong("PI(s): "), ', '.join(pis) or 'N/A']),
        html.P([html.Strong("Co-PI(s): "), ', '.join(copis) or 'N/A']),
        html.P([html.Strong("Grant Description: "), grant.description or ''])
    ], className="mb-4")


    # 2) Personnel List (name, role, email)
    model_map = {
        "PI": PI, "Co-PI": CoPI,
        "UI Professional Staff": ProfessionalStaff,
        "Postdoc": Postdoc, "GRA": GRA,
        "Undergrad": Undergrad, "Temp Help": TempHelp
    }
    people = session.query(distinct(GrantPersonnel.name),
                           GrantPersonnel.position) \
                   .filter_by(grant_id=grant_id).all()
    personnel_rows = []
    for name, position in people:
        mdl = model_map.get(position)
        email = session.query(mdl.email).filter_by(full_name=name).scalar() if mdl else ""
        personnel_rows.append({
            "Name": name,
            "Role": position,
            "Email": email or "—"
        })
    personnel_table = dash_table.DataTable(
        columns=[
            {"name": "Name", "id": "Name"},
            {"name": "Role", "id": "Role"},
            {"name": "Email", "id": "Email"},
        ],
        data=personnel_rows,
        style_table={"overflowX": "auto"},
        style_header={"fontWeight": "bold"},
        style_cell={"textAlign": "left"}
    )

    # 3) Personnel Hours by Year
    n_years = grant.duration
    # Build a lookup of summed hours per (name, position, year)
    sums = session.query(
        GrantPersonnel.name, GrantPersonnel.position, GrantPersonnel.year,
        func.coalesce(func.sum(GrantPersonnel.estimated_hours), 0).label("hrs")
    ).filter_by(grant_id=grant_id) \
     .group_by(GrantPersonnel.name, GrantPersonnel.position, GrantPersonnel.year) \
     .all()
    hours_map = {
        (r.name, r.position, r.year): float(r.hrs)
        for r in sums
    }

    # Build columns and rows
    hour_cols = [{"name": c, "id": c} for c in ["Name", "Role"]] + [
        {"name": f"Year {i+1}", "id": f"Year {i+1}"} for i in range(n_years)
    ]
    hour_rows = []
    for name, position in people:
        row = {"Name": name, "Role": position}
        for i in range(n_years):
            year = grant.start_date.year + i
            row[f"Year {i+1}"] = hours_map.get((name, position, year), 0.0)
        hour_rows.append(row)

    hours_table = dash_table.DataTable(
        columns=hour_cols,
        data=hour_rows,
        style_table={"overflowX": "auto"},
        style_header={"fontWeight": "bold"},
        style_cell={"textAlign": "center"}
    )



    # travel section
    # --- 4) Travel Itineraries ---
    travel_records = (
        session.query(GrantTravel)
               .filter_by(grant_id=grant_id)
               .all()
    )
    travel_rows = []
    for tr in travel_records:
        itin = tr.itinerary  # one-to-one
        travel_rows.append({
            "Type": tr.travel_type or "—",
            "Year": tr.year or "—",
            "Name": tr.name or "—",
            "Description": tr.description or "—",
            "Departure": itin.departure_date.strftime("%Y-%m-%d") if itin and itin.departure_date else "—",
            "Arrival":   itin.arrival_date.strftime("%Y-%m-%d")   if itin and itin.arrival_date   else "—",
            "Flight Cost": float(itin.flight_cost)   if itin and itin.flight_cost   else 0.0,
            "Days Stay":   itin.days_stay             if itin and itin.days_stay         else 0,
            "Food/Day":    float(itin.per_day_food_lodging)    if itin and itin.per_day_food_lodging    else 0.0,
            "Transp/Day":  float(itin.per_day_transportation)  if itin and itin.per_day_transportation  else 0.0,
        })

    travel_table = dash_table.DataTable(
        columns=[
            {"name": col, "id": col}
            for col in ["Type", "Year", "Name", "Description",
                        "Departure", "Arrival",
                        "Flight Cost", "Days Stay", "Food/Day", "Transp/Day"]
        ],
        data=travel_rows,
        style_table={"overflowX": "auto"},
        style_header={"fontWeight": "bold"},
        style_cell={"textAlign": "center", "whiteSpace": "normal"}
    )




    # material section
    # --- 5) Material Expenses ---

    mat_query = (
        session.query(
            GrantMaterial,
            ExpenseCategory.name.label("Category"),
            ExpenseSubcategory.name.label("Subcategory")
        )
        .join(ExpenseCategory,    GrantMaterial.category_id    == ExpenseCategory.id)
        .join(ExpenseSubcategory, GrantMaterial.subcategory_id == ExpenseSubcategory.id)
        .filter(GrantMaterial.grant_id == grant_id)
        .all()
    )

    mat_rows = []
    for gm, cat_name, subcat_name in mat_query:
        mat_rows.append({
            "Category":    cat_name,
            "Subcategory": subcat_name,
            "Year":        gm.year,
            "Description": gm.description or "",
            "Cost":        float(gm.cost or 0)
        })

    materials_table = dash_table.DataTable(
        columns=[{"name": c, "id": c} for c in ["Category", "Subcategory", "Year", "Description", "Cost"]],
        data=mat_rows,
        style_table={"overflowX": "auto"},
        style_header={"fontWeight": "bold"},
        style_cell={"textAlign": "center"}
    )


    session.close()

    modal_body = html.Div([
        grant_info,

        dbc.Row(
            dbc.Col(
                html.H6("Personnel List", className="text-center p-2"),
                width=12,
                className="bg-light rounded mb-2"
            )
        ),
        personnel_table,

        dbc.Row(
            dbc.Col(
                html.H6("Personnel Hours", className="text-center p-2"),
                width=12,
                className="bg-light rounded mb-2"
            )
        ),
        hours_table, 
        dbc.Row(
            dbc.Col(
                html.H6("Travel Information", className="text-center p-2"),
                width=12,
                className="bg-light rounded mb-2"
            )
        ),
        travel_table, 
        dbc.Row(
            dbc.Col(
                html.H6("Materials", className="text-center p-2"),
                width=12,
                className="bg-light rounded mb-2"
            )
        ),
        materials_table,
    ])
    return True, modal_body, grant_id


################################################
# Callback to handle the download button for PDF Dwonload
################################################
# helper function & style for wrapped cells
# define once:
USABLE_WIDTH = letter[0] - 30 - 30  # page width minus left+right margins

# helper function & style for wrapped cells
styles = getSampleStyleSheet()
wrap_style = ParagraphStyle(
    "wrap", parent=styles["BodyText"],
    wordWrap="CJK", fontSize=10, leading=12
)

def make_wrapped_table(data_rows, col_names, col_widths=None):
    """
    Builds a Table of Paragraphs (so text wraps).
    col_widths: list of widths (in points) or None to auto-divide evenly.
    """
    n = len(col_names)
    if col_widths is None:
        cw = USABLE_WIDTH / n
        col_widths = [cw] * n

    # header row as bold paragraphs
    hdr = [Paragraph(f"<b>{c}</b>", wrap_style) for c in col_names]
    tbl_data = [hdr]

    for row in data_rows:
        cells = []
        if isinstance(row, dict):
            vals = [row.get(c, "") for c in col_names]
        else:
            vals = list(row)
        for v in vals:
            cells.append(Paragraph(str(v), wrap_style))
        tbl_data.append(cells)

    tbl = Table(tbl_data, colWidths=col_widths)
    tbl.setStyle(TableStyle([
        ("GRID",   (0, 0), (-1, -1), 0.5, colors.black),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    return tbl

# --- your callback for PDF download ---
@callback(
    Output("download-pdf", "data"),
    Input("download-modal-pdf-btn", "n_clicks"),
    State("view-grant-id-store", "data"),
    prevent_initial_call=True
)
def download_pdf(n_clicks, grant_id):
    if not grant_id or n_clicks is None:
        raise PreventUpdate

    session = get_db_session()
    grant = session.query(Grant).get(grant_id)

    # 1) Grant Info
    pis = [r[0] for r in session.query(distinct(GrantPersonnel.name))
                          .filter_by(grant_id=grant_id, position="PI").all()]
    copis = [r[0] for r in session.query(distinct(GrantPersonnel.name))
                             .filter_by(grant_id=grant_id, position="Co-PI").all()]

    # 2) Personnel List
    model_map = {
        "PI": PI, "Co-PI": CoPI,
        "UI Professional Staff": ProfessionalStaff,
        "Postdoc": Postdoc, "GRA": GRA,
        "Undergrad": Undergrad, "Temp Help": TempHelp
    }
    people = session.query(distinct(GrantPersonnel.name), GrantPersonnel.position) \
                   .filter_by(grant_id=grant_id).all()
    personnel_rows = []
    for name, position in people:
        mdl = model_map.get(position)
        email = session.query(mdl.email).filter_by(full_name=name).scalar() if mdl else ""
        personnel_rows.append({"Name": name, "Role": position, "Email": email or "—"})

    # 3) Personnel Hours by Year
    n_years = grant.duration
    sums = session.query(
        GrantPersonnel.name, GrantPersonnel.position, GrantPersonnel.year,
        func.coalesce(func.sum(GrantPersonnel.estimated_hours), 0).label("hrs")
    ).filter_by(grant_id=grant_id) \
     .group_by(GrantPersonnel.name, GrantPersonnel.position, GrantPersonnel.year) \
     .all()
    hours_map = {(r.name, r.position, r.year): float(r.hrs) for r in sums}
    hour_rows = []
    for name, position in people:
        row = {"Name": name, "Role": position}
        for i in range(n_years):
            year = grant.start_date.year + i
            row[f"Year {i+1}"] = hours_map.get((name, position, year), 0.0)
        hour_rows.append(row)

    # 4) Travel Information
    travel_records = session.query(GrantTravel).filter_by(grant_id=grant_id).all()
    travel_rows = []
    for tr in travel_records:
        itin = tr.itinerary
        travel_rows.append({
            "Type": tr.travel_type or "—",
            "Year": tr.year or "—",
            "Name": tr.name or "—",
            "Description": tr.description or "—",
            "Departure": itin.departure_date.strftime("%Y-%m-%d") if itin and itin.departure_date else "—",
            "Arrival":   itin.arrival_date.strftime("%Y-%m-%d")   if itin and itin.arrival_date   else "—",
            "Flight Cost": float(itin.flight_cost)   if itin and itin.flight_cost   else 0.0,
            "Days Stay":   itin.days_stay             if itin and itin.days_stay         else 0,
            "Food/Day":    float(itin.per_day_food_lodging)    if itin and itin.per_day_food_lodging    else 0.0,
            "Transp/Day":  float(itin.per_day_transportation)  if itin and itin.per_day_transportation  else 0.0,
        })

    # 5) Materials
    mat_query = (
        session.query(GrantMaterial, ExpenseCategory.name.label("Category"), ExpenseSubcategory.name.label("Subcategory"))
               .join(ExpenseCategory,    GrantMaterial.category_id    == ExpenseCategory.id)
               .join(ExpenseSubcategory, GrantMaterial.subcategory_id == ExpenseSubcategory.id)
               .filter(GrantMaterial.grant_id == grant_id)
               .all()
    )
    mat_rows = []
    for gm, cat_name, subcat_name in mat_query:
        mat_rows.append({
            "Category":    cat_name,
            "Subcategory": subcat_name,
            "Year":        gm.year or "—",
            "Description": gm.description or "",
            "Cost":        float(gm.cost or 0)
        })

    session.close()

    # --- Build PDF ---
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=18)
    usable_width = letter[0] - doc.leftMargin - doc.rightMargin
    elems = []

    elems.append(Paragraph(f"Grant Report: {grant.title}", styles["Title"]))
    elems.append(Spacer(1, 12))
    elems.append(Paragraph(f"<b>Funding Agency:</b> {grant.funding_agency}", styles["Normal"]))
    elems.append(Paragraph(f"<b>Period:</b> {grant.start_date} → {grant.end_date}", styles["Normal"]))
    elems.append(Paragraph(f"<b>Duration:</b> {grant.duration} years", styles["Normal"]))
    # new fields
    elems.append(Paragraph(f"<b>PI(s):</b> {', '.join(pis) if pis else 'N/A'}", styles["Normal"]))
    elems.append(Paragraph(f"<b>Co-PI(s):</b> {', '.join(copis) if copis else 'N/A'}", styles["Normal"]))
    elems.append(Paragraph(f"<b>Grant Description:</b> {grant.description or 'N/A'}", styles["Normal"]))

    elems.append(Spacer(1, 12))

    # Personnel List
    elems.append(Paragraph("Personnel List", styles["Heading2"]))
    elems.append(make_wrapped_table(personnel_rows, ["Name", "Role", "Email"]))
    elems.append(Spacer(1, 12))

    # Personnel Hours
    hour_cols = ["Name", "Role"] + [f"Year {i+1}" for i in range(n_years)]
    elems.append(Paragraph("Personnel Hours", styles["Heading2"]))
    elems.append(make_wrapped_table(hour_rows, hour_cols))
    elems.append(Spacer(1, 12))

    # Travel Information
    travel_cols = ["Type","Year","Name","Description","Departure","Arrival","Flight Cost","Days Stay","Food/Day","Transp/Day"]
    # custom widths to give Description more room
    w = usable_width
    travel_widths = [0.08*w,0.06*w,0.08*w,0.30*w,0.08*w,0.08*w,0.10*w,0.06*w,0.08*w,0.08*w]
    elems.append(Paragraph("Travel Information", styles["Heading2"]))
    elems.append(make_wrapped_table(travel_rows, travel_cols, travel_widths))
    elems.append(Spacer(1, 12))

    # Materials
    mat_cols = ["Category","Subcategory","Year","Description","Cost"]
    w = usable_width
    mat_widths = [0.15*w,0.15*w,0.08*w,0.40*w,0.22*w]
    elems.append(Paragraph("Materials", styles["Heading2"]))
    elems.append(make_wrapped_table(mat_rows, mat_cols, mat_widths))

    doc.build(elems)
    pdf = buf.getvalue()
    buf.close()
    return dcc.send_bytes(pdf, filename=f"grant_{grant_id}.pdf")



################################################
# Callback to handle the download button for Excel
################################################
@callback(
    Output("download-excel-modal", "data"),
    Input("download-modal-excel-btn", "n_clicks"),
    State("view-grant-id-store", "data"),
    prevent_initial_call=True,
)
def download_excel(n_clicks, grant_id):
    if not grant_id or not n_clicks:
        raise PreventUpdate

    session = get_db_session()
    grant = session.query(Grant).get(grant_id)

    # --- 1) Grant Info as a one‐row DataFrame ---
    pis = [r[0] for r in session
       .query(distinct(GrantPersonnel.name))
       .filter_by(grant_id=grant_id, position="PI")
       .all()]
    copis = [r[0] for r in session
       .query(distinct(GrantPersonnel.name))
       .filter_by(grant_id=grant_id, position="Co-PI")
       .all()]

    info_df = pd.DataFrame([{
        "Title": grant.title,
        "Funding Agency": grant.funding_agency,
        "Period": f"{grant.start_date} → {grant.end_date}",
        "Duration (years)": grant.duration,
        "PI(s)": ", ".join(pis) or "N/A",
        "Co-PI(s)": ", ".join(copis) or "N/A",
        "Description": grant.description or ""
    }])

    # --- 2) Personnel List ---
    model_map = {
        "PI": PI, "Co-PI": CoPI,
        "UI Professional Staff": ProfessionalStaff,
        "Postdoc": Postdoc, "GRA": GRA,
        "Undergrad": Undergrad, "Temp Help": TempHelp
    }
    people = session.query(distinct(GrantPersonnel.name), GrantPersonnel.position) \
                   .filter_by(grant_id=grant_id).all()
    per_rows = []
    for name, pos in people:
        mdl = model_map.get(pos)
        email = session.query(mdl.email).filter_by(full_name=name).scalar() if mdl else ""
        per_rows.append({"Name": name, "Role": pos, "Email": email or "—"})
    per_df = pd.DataFrame(per_rows)

    # --- 3) Personnel Hours by Year ---
    n_years = grant.duration
    sums = session.query(
        GrantPersonnel.name,
        GrantPersonnel.position,
        GrantPersonnel.year,
        func.coalesce(func.sum(GrantPersonnel.estimated_hours), 0).label("hrs")
    ).filter_by(grant_id=grant_id) \
     .group_by(GrantPersonnel.name, GrantPersonnel.position, GrantPersonnel.year) \
     .all()
    hrs_map = {(r.name, r.position, r.year): float(r.hrs) for r in sums}
    hrs_rows = []
    for name, pos in people:
        row = {"Name": name, "Role": pos}
        for i in range(n_years):
            year = grant.start_date.year + i
            row[f"Year {i+1}"] = hrs_map.get((name, pos, year), 0.0)
        hrs_rows.append(row)
    hrs_df = pd.DataFrame(hrs_rows)

    # --- 4) Travel Itineraries ---
    trav_recs = session.query(GrantTravel).filter_by(grant_id=grant_id).all()
    trav_rows = []
    for tr in trav_recs:
        it = tr.itinerary
        trav_rows.append({
            "Type": tr.travel_type or "—",
            "Year": tr.year or "—",
            "Name": tr.name or "—",
            "Description": tr.description or "—",
            "Departure": it.departure_date.strftime("%Y-%m-%d") if it and it.departure_date else "—",
            "Arrival":   it.arrival_date.strftime("%Y-%m-%d")   if it and it.arrival_date   else "—",
            "Flight Cost": float(it.flight_cost) if it and it.flight_cost else 0.0,
            "Days Stay":   it.days_stay     if it and it.days_stay     else 0,
            "Food/Day":    float(it.per_day_food_lodging)    if it and it.per_day_food_lodging    else 0.0,
            "Transp/Day":  float(it.per_day_transportation)  if it and it.per_day_transportation  else 0.0,
        })
    trav_df = pd.DataFrame(trav_rows)

    # --- 5) Materials ---
    mat_q = (
        session.query(
            GrantMaterial,
            ExpenseCategory.name.label("Category"),
            ExpenseSubcategory.name.label("Subcategory")
        )
        .join(ExpenseCategory,    GrantMaterial.category_id    == ExpenseCategory.id)
        .join(ExpenseSubcategory, GrantMaterial.subcategory_id == ExpenseSubcategory.id)
        .filter(GrantMaterial.grant_id == grant_id)
        .all()
    )
    mat_rows = []
    for gm, cat, sub in mat_q:
        mat_rows.append({
            "Category":    cat,
            "Subcategory": sub,
            "Year":        gm.year or "—",
            "Description": gm.description or "",
            "Cost":        float(gm.cost or 0)
        })
    mat_df = pd.DataFrame(mat_rows)

    session.close()

    # --- Write everything into one sheet, with spacing between sections ---
    buf = io.BytesIO()
    # Section‐title format
    
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        workbook  = writer.book
        worksheet = workbook.add_worksheet("Report")
        writer.sheets["Report"] = worksheet

        section_fmt = workbook.add_format({
            "bold": True,
            "font_size": 12,
            "bg_color": "#4F81BD",   # blue
            "font_color": "white",
            "border": 1
        })

        # Header‐row format
        header_fmt = workbook.add_format({
            "bold": True,
            "bg_color": "#D9E1F2",   # light blue
            "border": 1
        })

        # Alternate‐row format
        alt_fmt = workbook.add_format({
            "bg_color": "#F2F2F2"    # very light grey
        })

        # helper for writing a dataframe at a given row
        def sheet_df(df, start_row, title):
            # 1) Section title
            worksheet.write(start_row, 0, title, section_fmt)

            # 2) Manually write & format the header row one row below the title
            for col_num, col_name in enumerate(df.columns):
                worksheet.write(start_row + 1, col_num, col_name, header_fmt)

            # 3) Write the data starting two rows below the title (i.e. below your header)
            df.to_excel(
                writer,
                sheet_name="Report",
                index=False,
                header=False,
                startrow=start_row + 2
            )

            # 4) Apply alternating‐row colours
            for i in range(len(df)):
                row_idx = start_row + 2 + i
                if i % 2 == 1:
                    worksheet.set_row(row_idx, cell_format=alt_fmt)

            # 5) Return the next free row (+2 rows of spacing)
            return start_row + 2 + len(df) + 2



        row = 0
        row = sheet_df(info_df,  row, "Grant Info")
        row = sheet_df(per_df,   row, "Personnel List")
        row = sheet_df(hrs_df,   row, "Personnel Hours")
        row = sheet_df(trav_df,  row, "Travel Information")
        row = sheet_df(mat_df,   row, "Materials")

        # you can also auto‐adjust column widths here if desired...
        max_cols = max(
            len(info_df.columns),
            len(per_df.columns),
            len(hrs_df.columns),
            len(trav_df.columns),
            len(mat_df.columns)
        )
        worksheet.set_column(0, max_cols - 1, 23)

    buf.seek(0)
    return dcc.send_bytes(buf.read(), filename=f"grant_{grant_id}.xlsx")





@callback(
    Output('table-container', 'children', allow_duplicate=True),
    Input({'type': 'delete-btn', 'index': ALL}, 'n_clicks'),
    prevent_initial_call=True
)
def delete_grant(n_clicks_list):
    # Check if the delete button was clicked (ensure n_clicks_list is not empty)
    if not any(n_clicks_list):
        return dash.no_update  # No update if no button was clicked

    # Extract the button ID from the triggered property
    ctx = dash.callback_context
    if not ctx.triggered:
        return dash.no_update

    # Extract the grant ID
    button_id = ctx.triggered[0]['prop_id'].split('.')[0]
    grant_id = eval(button_id)['index']  # Extract the index (grant ID)

    # Get the database session
    session = get_db_session()

    try:
        # Ensure the current user is logged in
        if not current_user.is_authenticated:
            return html.Div("You must be logged in to delete a grant.", style={"color": "red"})

        # Fetch the grant by ID
        grant_to_delete = session.query(Grant).get(grant_id)

        # Ensure the current user is the owner of the grant
        if grant_to_delete.user_id != current_user.id:
            return html.Div("You do not have permission to delete this grant.", style={"color": "red"})

        # Delete the grant from the database
        session.delete(grant_to_delete)
        session.commit()

        # Refresh the table after deletion
        return display_grants(0)

    except Exception as err:
        return html.Div(f"Error: {str(err)}", style={"color": "red"})

    finally:
        session.close()



position_model_map = {
    "PI": PI,
    "Co-PI": CoPI,
    "UI Professional Staff": ProfessionalStaff,
    "Postdoc": Postdoc,
    "GRA": GRA,
    "Temp Help": TempHelp,
    "Undergrad": Undergrad
}

@callback(
    Output("download-excel", "data"),
    Input({'type': 'download-excel-btn', 'index': ALL}, "n_clicks"),
    prevent_initial_call=True
)
def download_excel(n_clicks_list):
    if not any(n_clicks_list):
        raise PreventUpdate

    triggered = ctx.triggered_id
    if not triggered:
        raise PreventUpdate

    try:
        grant_id = triggered['index']
        print("grant id is", grant_id)
        session = get_db_session()

        if not current_user.is_authenticated:
            raise PreventUpdate

        grant = session.query(Grant).get(grant_id)
        print("grant is", grant)
        # get the duration
        duration = session.query(Grant.duration).filter_by(id=grant_id).scalar()
        print("duration is", duration)

        funding_agency = session.query(Grant.funding_agency).filter_by(id=grant_id).scalar()
        print("funding agency is", funding_agency)

        if not grant or grant.user_id != current_user.id:
            raise PreventUpdate

        # Create Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Grant Budget"
        # set the row height
        for i in range(1, 101):
            ws.row_dimensions[i].height = 20
        # Set font size to 12 for all written cells
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:  # Only set font for non-empty cells
                    cell.font = Font(size=12)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        bold = Font(bold=True)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ######################################
        # Section 1: Grant Information
        ######################################
        # Row 1
        ws["A1"] = "Title:"
        ws["A1"].font = bold
        ws["B1"] = grant.title or ""
        ws["J1"] = "PI FTE"
        ws["J1"].font = bold
        ws["J1"].fill = yellow_fill
        ws["K1"] = 0.00
        ws["K1"].fill = yellow_fill

        # Merge B1:I1
        ws.merge_cells('B1:I1')
        # Merge B2:I2
        ws.merge_cells('B2:I2')
        # merge from B3:I3
        ws.merge_cells('B3:I3')

        # Row 2
        ws["A2"] = "Funding source:"
        ws["A2"].font = bold
        ws["B2"] = grant.funding_agency or ""

        # --- Get PI ---
        # There should only be ONE PI normally, so just use a simple query without distinct
        pi_obj = session.query(GrantPersonnel).filter_by(grant_id=grant.id, position="PI").first()
        pi_name = pi_obj.name if pi_obj else "N/A"

        # --- Get unique CoPIs ---
        copi_names = session.query(distinct(GrantPersonnel.name)) \
            .filter_by(grant_id=grant.id, position="Co-PI") \
            .all()

        # Flatten tuple result
        copi_names_list = [name[0] for name in copi_names]
        CoPIs_names = ", ".join(copi_names_list) if copi_names_list else "N/A"

        ws["A3"] = f"PIs: {pi_name}"
        ws["A3"].font = bold
        ws["B3"] = f"CoPIs: {CoPIs_names}"
        ws["B3"].font = bold
        ws["C3"].font = bold

    

        # Row 4
        ws["A4"] = "Project Start and End Dates:"
        ws["A4"].font = bold
        if grant.start_date and grant.end_date:
            ws["B4"] = f"{grant.start_date.strftime('%Y-%m-%d')}  -  {grant.end_date.strftime('%Y-%m-%d')}"

        # Adjust column widths
        ws.column_dimensions["A"].width = 39
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["J"].width = 32.5
        for col in range(4, 9):  # D to I
            ws.column_dimensions[chr(64+col)].width = 18



        ######################################
        # Section 2: Personnel Compensation
        ######################################
        ws.column_dimensions['C'].width = 11
        ws.column_dimensions['D'].width = 11
        ws.column_dimensions['E'].width = 11
        ws.column_dimensions['F'].width = 11
        ws.column_dimensions['G'].width = 11
        ws.column_dimensions['H'].width = 11
        ws.column_dimensions['I'].width = 11
        # Define a thin black border
        thin_border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        # Apply the border to all cells in Row 5 (C5, D5, E5, etc.)
        for i in range(1, 11):  
            cell = ws.cell(row=5, column=i)
            cell.border = thin_border
        header = ["Hourly rate at Start Date", "Y1", "Y2", "Y3", "Y4", "Y5", "Total", "Notes"]
        ws.row_dimensions[5].height = 32
        # Start writing from column 3 (which is Column C in Excel)
        for i, text in enumerate(header):
            cell = ws.cell(row=5, column=3 + i, value=text)  # 3 = column C
            cell.font = bold
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True )

    
        # === Row 6: Personnel Compensation Section Header ===   define styles
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        bold_underline_font = Font(bold=True, underline="single")
        underline_font = Font(underline="single")


        # Row 6: Personnel Compensation (apply fill across A6 to J6)
        ws["A6"] = "Personnel Compensation"
        ws["B6"] = "Y1 Hours"
        ws["B6"].font = bold_underline_font
        ws["A6"].font = underline_font
        ws["A6"].alignment = Alignment(horizontal="left", vertical="center")
        

        # Fill each cell from A6 to J6 with gray (no merging)
        for col in range(1, 11):  # Columns A(1) to J(10)
            ws.cell(row=6, column=col).fill = gray_fill

        row = 7
        positions_to_include = ["PI", "Co-PI"]

        for position in positions_to_include:
            # Distinct personnel names
            people = session.query(distinct(GrantPersonnel.name)) \
                            .filter_by(grant_id=grant.id, position=position) \
                            .all()
            people_list = [p[0] for p in people]

            for person in people_list:
                # Estimated hours for first year
                first_record = session.query(GrantPersonnel) \
                    .filter_by(grant_id=grant.id, position=position, name=person) \
                    .order_by(asc(GrantPersonnel.year)) \
                    .first()

                ws[f"A{row}"] = position
                ws[f"B{row}"] = first_record.estimated_hours if first_record else "-"

                # Get hourly rate
                hourly_rate = None
                if position == "PI":
                    hourly_rate = session.query(PI.expected_hourly_salary).filter_by(full_name=person).scalar()
                elif position == "Co-PI":
                    hourly_rate = session.query(CoPI.expected_hourly_salary).filter_by(full_name=person).scalar()

                ws[f"C{row}"] = float(hourly_rate) if hourly_rate else "-"

                # Determine funding compensation table
                funding_agency = grant.funding_agency.lower()
                if "nsf" in funding_agency:
                    comp_table = NSFPersonnelCompensation
                elif "nih" in funding_agency:
                    comp_table = NIHPersonnelCompensation
                else:
                    comp_table = None

                rate_increase = session.query(comp_table).filter(comp_table.role.ilike(position)).first() if comp_table else None

                current_hourly_rate = float(hourly_rate)  # Start with the base hourly rate
                for i in range(grant.duration):
                    year = grant.start_date.year + i

                    # Get that year's hours
                    hours_record = session.query(GrantPersonnel.estimated_hours) \
                        .filter_by(grant_id=grant.id, position=position, name=person, year=year) \
                        .scalar()

                    col_letter = chr(68 + i)  # Column D = ASCII 68

                    if hours_record and current_hourly_rate:
                        if i > 0:  # Apply raise starting from Year 2
                            increase = getattr(rate_increase, f"y{i+1}_rate_increase", 0.0) if rate_increase else 0.0
                            current_hourly_rate = current_hourly_rate * (1.0 + float(increase))

                        total_expense = float(hours_record) * current_hourly_rate
                        ws[f"{col_letter}{row}"] = round(total_expense, 2)
                    else:
                        ws[f"{col_letter}{row}"] = "-"

                    # fill the extra columns with "-"
                    max_duration = 5
                    if grant.duration < max_duration:
                        for i in range(grant.duration, max_duration):
                            col_letter = chr(68 + i)
                            cell = ws[f"{col_letter}{row}"]
                            cell.value = "-"
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Compute total for valid project years
                    total = 0.0
                    for i in range(grant.duration):
                        col_letter = chr(68 + i)  # Column D is ASCII 68
                        cell_value = ws[f"{col_letter}{row}"].value
                        if isinstance(cell_value, (int, float)):
                            total += cell_value

                    # Write the total in Column I (ASCII 73)
                    ws["I" + str(row)] = round(total, 2)
                    ws["I" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
                row += 1

        # === Write 'Other personnel' header ===
        row +=1
        ws[f"A{row}"] = "Other personnel"
        ws[f"A{row}"].font = underline_font  # Or use bold_underline_font if you want both
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 11):  # Columns A(1) to J(10)
            ws.cell(row=row, column=col).fill = gray_fill
        


        # for other staffs
        # === Handle all other personnel positions ===
        row +=1
        all_positions = session.query(distinct(GrantPersonnel.position)) \
            .filter(GrantPersonnel.grant_id == grant.id) \
            .all()
        other_positions = [pos[0] for pos in all_positions if pos[0] not in ["PI", "Co-PI"]]
        print("other positions are", other_positions)

        for position in other_positions:
            print("position is : ", position)
            # Distinct people for this position
            people = session.query(distinct(GrantPersonnel.name)) \
                            .filter_by(grant_id=grant.id, position=position) \
                            .all()
            people_list = [p[0] for p in people]
            print("people list is : ", people_list)

            for person in people_list:
                print("pserson is : ", person)
                # Estimated hours for first year
                first_record = session.query(GrantPersonnel) \
                    .filter_by(grant_id=grant.id, position=position, name=person) \
                    .order_by(asc(GrantPersonnel.year)) \
                    .first()

                ws[f"A{row}"] = position
                ws[f"B{row}"] = first_record.estimated_hours if first_record else "-"

                # Get hourly rate from general staff table (assume single table for now)
                model_class = position_model_map.get(position)
                print("model class is : ", model_class)
                print("position is : ", position)
                hourly_rate = None
                if model_class:
                    hourly_rate = session.query(model_class.expected_hourly_salary).filter_by(full_name=person).scalar()


                ws[f"C{row}"] = float(hourly_rate) if hourly_rate else "-"

                # Determine compensation table
                funding_agency = grant.funding_agency.lower()
                if "nsf" in funding_agency:
                    comp_table = NSFPersonnelCompensation
                elif "nih" in funding_agency:
                    comp_table = NIHPersonnelCompensation
                else:
                    comp_table = None

                rate_increase = session.query(comp_table).filter(comp_table.role.ilike(position)).first() if comp_table else None

                current_hourly_rate = float(hourly_rate) if hourly_rate else 0.0
                for i in range(grant.duration):
                    year = grant.start_date.year + i
                    col_letter = chr(68 + i)

                    hours_record = session.query(GrantPersonnel.estimated_hours) \
                        .filter_by(grant_id=grant.id, position=position, name=person, year=year) \
                        .scalar()

                    if hours_record and current_hourly_rate:
                        if i > 0:
                            increase = getattr(rate_increase, f"y{i+1}_rate_increase", 0.0) if rate_increase else 0.0
                            current_hourly_rate *= (1.0 + float(increase))

                        total_expense = float(hours_record) * current_hourly_rate
                        ws[f"{col_letter}{row}"] = round(total_expense, 2)
                    else:
                        ws[f"{col_letter}{row}"] = "-"

                # Fill extra unused years with "-"
                max_duration = 5
                if grant.duration < max_duration:
                    for i in range(grant.duration, max_duration):
                        col_letter = chr(68 + i)
                        cell = ws[f"{col_letter}{row}"]
                        cell.value = "-"
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                # Compute total for valid years
                total = 0.0
                for i in range(grant.duration):
                    col_letter = chr(68 + i)
                    val = ws[f"{col_letter}{row}"].value
                    if isinstance(val, (int, float)):
                        total += val

                ws[f"I{row}"] = round(total, 2)
                ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")
                row += 1




        # === Fringe Header Row ===
        row += 1
        ws[f"A{row}"] = "Fringe"
        ws[f"A{row}"].font = underline_font
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

        # Fill across A to J
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = gray_fill

        # FY Label
        first_year = grant.start_date.year
        second_year = first_year + 1
        ws[f"C{row}"] = f"FY{str(first_year)[-2:]}-{str(second_year)[-2:]} Fringe Rates"
        ws[f"C{row}"].font = Font(bold=True, underline='single', color="0000FF")
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center")

        # Select appropriate fringe table
        funding_agency = grant.funding_agency.lower()
        fringe_table = None
        if "nsf" in funding_agency:
            fringe_table = NSFFringeRate
        elif "nih" in funding_agency:
            fringe_table = NIHFringeRate

        # === FY Label in Column C ===
        first_year = grant.start_date.year
        second_year = first_year + 1
        ws[f"C{row}"] = f"FY{str(first_year)[-2:]}-{str(second_year)[-2:]} Fringe Rates"
        ws[f"C{row}"].font = Font(bold=True, underline='single', color="0000FF")
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center")

        # Determine which fringe rate table to use
        funding_agency = grant.funding_agency.lower()
        if "nsf" in funding_agency:
            fringe_table = NSFFringeRate
        elif "nih" in funding_agency:
            fringe_table = NIHFringeRate
        else:
            fringe_table = None


        # Mapping from positions to fringe categories
        fringe_category_map = {
            "PI": "Faculty",
            "Co-PI": "Faculty",
            "UI Professional Staff": "UI professional staff & Post Docs",
            "Postdoc": "UI professional staff & Post Docs",
            "GRA": "GRAs/UGrads",
            "Undergrad": "GRAs/UGrads",
            "Temp Help": "Temp Help"
        }

        # Build compensation lookup: { (role, year): total compensation }
        comp_lookup = {}
        for r in range(7, row):  # adjust as needed based on personnel section start
            position = ws[f"A{r}"].value
            category = fringe_category_map.get(position)
            if not category:
                continue
            for i in range(grant.duration):
                col_letter = chr(68 + i)  # D = 68
                comp = ws[f"{col_letter}{r}"].value
                if isinstance(comp, (int, float)):
                    comp_lookup[(category, i+1)] = comp_lookup.get((category, i+1), 0) + comp

        ## === Write Fringe Rows per Category ===
        roles = ["Faculty", "UI professional staff & Post Docs", "GRAs/UGrads", "Temp Help"]
        for role in roles:
            row += 1
            ws[f"A{row}"] = role
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

            # Get Year 1 fringe rate and display it in column C
            first_year_rate = session.query(fringe_table.fringe_rate).filter_by(role=role, year=1).scalar() if fringe_table else None
            ws[f"C{row}"] = f"{float(first_year_rate):.1f}%" if first_year_rate is not None else "-"
            ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center")

            total_fringe = 0.0
            for i in range(grant.duration):
                year = i + 1
                comp = comp_lookup.get((role, year), 0.0)
                rate = session.query(fringe_table.fringe_rate).filter_by(role=role, year=year).scalar() if fringe_table else 0.0
                fringe_amt = comp * float(rate) / 100 if rate else 0.0
                total_fringe += fringe_amt

                col_letter = chr(68 + i)  # D, E, F...
                ws[f"{col_letter}{row}"] = round(fringe_amt, 2) if fringe_amt else "-"
                ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Fill unused years with "-"
            for i in range(grant.duration, 5):
                col_letter = chr(68 + i)
                ws[f"{col_letter}{row}"] = "-"
                ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Write total in Column I
            # ws[f"I{row}"] = round(total_fringe, 2)
            ws[f"I{row}"] = round(total_fringe, 2) if total_fringe > 0 else "-"
            ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")
        
        
        
        
        # === Leave one empty row after Fringe ===
        # === Leave one empty row after Fringe ===
        row += 1

        # === Equipment >$5000.00 Section Header ===
        row += 1
        ws[f"A{row}"] = "Equipment >$5000.00"
        ws[f"A{row}"].font = underline_font
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

        # Gray fill from A to J
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = gray_fill

        # # === Column titles below header ===
        # row += 1
        # ws[f"A{row}"] = "Item Description"
        # ws[f"A{row}"].font = bold
        # ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

        ws[f"I{row}"] = "Total"
        ws[f"I{row}"].font = bold
        ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")

        # Fill background for row
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = gray_fill

        # === Query Equipment >$5000.00 Items ===
        equipment_items = (
            session.query(GrantMaterial)
            .join(ExpenseSubcategory, GrantMaterial.subcategory_id == ExpenseSubcategory.id)
            .filter(
                GrantMaterial.grant_id == grant.id,
                ExpenseSubcategory.name == "Equipment >5K"
            )
            .all()
        )

        # === Render Each Equipment Item ===
        for item in equipment_items:
            row += 1
            ws[f"A{row}"] = item.description or "-"
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

            total_cost = 0.0
            item_year = item.year
            base_year = grant.start_date.year
            offset = item_year - base_year

            for i in range(grant.duration):  # Loop from Y1 to Y(n)
                col_letter = chr(68 + i)  # D = 68
                if offset == i:
                    ws[f"{col_letter}{row}"] = round(item.cost, 2)
                    total_cost = item.cost
                else:
                    ws[f"{col_letter}{row}"] = "-"
                ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Fill unused years (up to Y5)
            for i in range(grant.duration, 5):
                col_letter = chr(68 + i)
                ws[f"{col_letter}{row}"] = "-"
                ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Total in Column I
            ws[f"I{row}"] = round(total_cost, 2) if total_cost > 0 else "-"
            ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")







        # === Leave one empty row after Equipment section ===
        row += 1

        # === Travel Section Header ===
        row += 1
        ws[f"A{row}"] = "Travel"
        ws[f"A{row}"].font = underline_font
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

        # Gray fill across A to J
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = gray_fill
        
        # === Query all travel records for this grant ===
        travel_records = (
            session.query(GrantTravel)
            .filter_by(grant_id=grant.id)
            .all()
        )

        # Create mapping: {"Domestic": {year: total}, "International": {year: total}}
        travel_summary = {"Domestic": {}, "International": {}}

        for travel in travel_records:
            total_cost = 0.0
            itineraries = session.query(TravelItinerary).filter_by(travel_id=travel.id).all()

            for itin in itineraries:
                days = itin.days_stay or 0
                food = float(itin.per_day_food_lodging or 0)
                transport = float(itin.per_day_transportation or 0)
                flight = float(itin.flight_cost or 0)

                total_cost += flight + days * (food + transport)


            if travel.travel_type in travel_summary:
                year_total = travel_summary[travel.travel_type].get(travel.year, 0.0)
                travel_summary[travel.travel_type][travel.year] = year_total + total_cost

        # === Render Domestic and International Rows ===
        for travel_type in ["Domestic", "International"]:
            row += 1
            ws[f"A{row}"] = f"{travel_type} Travel"
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

            total_for_type = 0.0
            year_costs = travel_summary[travel_type]

            base_year = grant.start_date.year

            for i in range(grant.duration):
                year = base_year + i
                col_letter = chr(68 + i)  # D = 68
                cost = year_costs.get(year)

                if cost:
                    ws[f"{col_letter}{row}"] = round(cost, 2)
                    total_for_type += cost
                else:
                    ws[f"{col_letter}{row}"] = "-"

                ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Fill unused years
            for i in range(grant.duration, 5):
                col_letter = chr(68 + i)
                ws[f"{col_letter}{row}"] = "-"
                ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Total in Column I
            ws[f"I{row}"] = round(total_for_type, 2) if total_for_type > 0 else "-"
            ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")




        # particpipant Support cost - nsf only
        # === Participant Support Costs (NSF ONLY) Section ===
        if "nsf" in grant.funding_agency.lower():
            # Leave an empty row before the section
            row += 1

            # Section header
            row += 1
            ws[f"A{row}"] = "Participant Support Costs (NSF ONLY)"
            ws[f"A{row}"].font = underline_font
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

            # Fill across A–J with gray
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = gray_fill

            # Placeholder rows (2 rows)
            for label in ["Participant Stipends", "Participant Travel"]:
                row += 1
                ws[f"A{row}"] = ""
                ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

                # Y1 to Y5 and Total as "-"
                for i in range(5):  # Y1 to Y5
                    col_letter = chr(68 + i)  # D = 68
                    ws[f"{col_letter}{row}"] = "-"
                    ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")

                # Total column I
                ws[f"I{row}"] = "-"
                ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")






        # Other directs cost
        # === Leave one empty row after previous section ===
        row += 1

        # Header row
        row += 1
        ws[f"A{row}"] = "Other Direct Costs"
        ws[f"A{row}"].font = underline_font
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = gray_fill

        # === Query all subcategories under 'Other Direct Costs' (category_id = 4) ===
        subcategories = session.query(ExpenseSubcategory).filter_by(category_id=4).all()
        subcategory_ids = [s.id for s in subcategories]
        subcategory_names = {s.id: s.name for s in subcategories}

        # Get all materials linked to this grant that fall under "Other Direct Costs"
        materials = (
            session.query(GrantMaterial)
            .filter(GrantMaterial.grant_id == grant.id, GrantMaterial.subcategory_id.in_(subcategory_ids))
            .all()
        )

        # Group materials by (subcategory_id, year)
        material_lookup = {}
        for m in materials:
            key = (m.subcategory_id, m.year)
            material_lookup[key] = material_lookup.get(key, 0) + float(m.cost)

        # Check if this grant includes a GRA
        has_gra = session.query(GrantPersonnel).filter_by(grant_id=grant.id, position="GRA").first() is not None

        # Get graduate cost info (only if GRA exists)
        tuition_rows = []
        if has_gra:
            grad_cost = session.query(GraduateStudentCost).first()
            if grad_cost:
                base = float(grad_cost.base_tuition_per_semester) * grad_cost.num_semesters_per_year
                summer = float(grad_cost.summer_credit_cost)
                insurance = float(grad_cost.health_insurance_cost)
                increase = float(grad_cost.annual_increase_percent)

                for i in range(grant.duration):
                    factor = (1 + increase / 100) ** i
                    yearly_total = round((base + summer + insurance) * factor, 2)
                    tuition_rows.append(yearly_total)

        # === Render one row per subcategory ===
        for sub in subcategories:
            row += 1
            ws[f"A{row}"] = sub.name
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

            total = 0.0

            # Special handling for graduate tuition
            if sub.name == "Grad Student Tuition & Health Insurance" and has_gra and tuition_rows:
                for i in range(grant.duration):
                    col_letter = chr(68 + i)  # D = 68
                    val = tuition_rows[i]
                    ws[f"{col_letter}{row}"] = val
                    ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")
                    total += val
            else:
                # Standard material costs by subcategory and year
                for i in range(grant.duration):
                    year = grant.start_date.year + i
                    col_letter = chr(68 + i)
                    cost = material_lookup.get((sub.id, year))
                    if cost:
                        ws[f"{col_letter}{row}"] = round(cost, 2)
                        total += cost
                    else:
                        ws[f"{col_letter}{row}"] = "-"
                    ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Fill remaining year columns (if duration < 5)
            for i in range(grant.duration, 5):
                col_letter = chr(68 + i)
                ws[f"{col_letter}{row}"] = "-"
                ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Total column I
            ws[f"I{row}"] = round(total, 2) if total > 0 else "-"
            ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")





        # === Leave one empty row after previous section ===
        row += 1

        # === Consortia/Subawards Section Header ===
        row += 1
        ws[f"A{row}"] = "Consortia/Subawards"
        ws[f"A{row}"].font = underline_font
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

        # Gray fill from A to J
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = gray_fill

        # === Subaward Placeholder Rows ===
        for label in ["Sub award 1", "Sub award 2"]:
            row += 1
            ws[f"A{row}"] = label
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

            for i in range(5):  # Y1–Y5
                col_letter = chr(68 + i)
                ws[f"{col_letter}{row}"] = "-"
                ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Total column I
            ws[f"I{row}"] = "-"
            ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")




        # Total Direct Costs
        # === Leave one row after previous section ===
        row += 1

        # === Total Direct Cost Header Row ===
        row += 1
        ws[f"A{row}"] = "Total Direct Cost"
        ws[f"A{row}"].font = bold
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

        # Fill Y1–Y5 and Total columns with 0
        for i in range(5):  # Columns D–H
            col_letter = chr(68 + i)
            ws[f"{col_letter}{row}"] = 0
            ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")

        # Total Column I
        ws["I" + str(row)] = 0
        ws["I" + str(row)].alignment = Alignment(horizontal="center", vertical="center")

        # Gray fill for header
        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = gray_fill

        # === Placeholder Rows ===
        placeholders = [
            "Back out GRA T&F",
            "Back out capital EQ",
            "Back out subawards totals",
            "Sub award 1 1st $25k",
            "Sub award 2 1st $25k",
            "Modified Total Direct Costs"
        ]
        # Define light gray fill
        light_gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Define all rows, mark the ones needing light gray fill
        rows_with_light_gray = [
            "Back out subawards totals",
            "Sub award 1 1st $25k",
            "Sub award 2 1st $25k"
        ]
        bold_rows = [
            "Modified Total Direct Costs",
            "Indirect Costs"
        ]


        for label in placeholders:
            row += 1
            ws[f"A{row}"] = label
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

            # Bold the row label if needed
            if label in bold_rows:
                ws[f"A{row}"].font = bold

            for i in range(5):  # Y1–Y5
                col_letter = chr(68 + i)
                ws[f"{col_letter}{row}"] = "-"
                ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Total
            ws[f"I{row}"] = "-"
            ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Apply light gray fill if needed
            if label in rows_with_light_gray:
                for col in range(1, 11):
                    ws.cell(row=row, column=col).fill = light_gray_fill

        # === Indirect Costs row ===
        row += 1
        ws[f"A{row}"] = "Indirect Costs"
        ws[f"A{row}"].font = bold
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"C{row}"] = "50.0%"
        ws[f"C{row}"].font = Font(bold=True)
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center")
        for i in range(5):  # Y1–Y5
            col_letter = chr(68 + i)
            ws[f"{col_letter}{row}"] = "-"
            ws[f"{col_letter}{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"I{row}"] = "-"
        ws[f"I{row}"].alignment = Alignment(horizontal="center", vertical="center")

        # === Total Project Cost row ===
        row += 1
        ws[f"A{row}"] = "Total Project Cost"
        ws[f"A{row}"].font = bold
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")

        # Styling for A–I
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.fill = gray_fill
            cell.font = bold
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000")
            )

        # Leave column C blank
        ws[f"C{row}"] = ""

        project_cost_start_row = 7               # Update to your actual start row
        project_cost_end_row = row - 1           # One row before "Total Project Cost"

        # === Dynamically calculate Y1–Y5 column totals using SUMIF ===
        for i in range(5):  # Columns D to H
            col_letter = chr(68 + i)  # D = 68
            cell_range = f"{col_letter}{project_cost_start_row}:{col_letter}{project_cost_end_row}"
            ws[f"{col_letter}{row}"] = f'=SUMIF({cell_range},">0")'

        # === Total column I: Sum of D–H in this row ===
        ws[f"I{row}"] = f"=SUM(D{row}:H{row})"


       



        

        # Save to in-memory stream
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return dcc.send_bytes(
            output.getvalue(),
            filename=f"grant_{grant_id}_budget.xlsx"
        )

    except Exception as err:
        print(f"Error while downloading Excel: {err}")
        raise PreventUpdate

    finally:
        try:
            session.close()
        except:
            pass



# callack for the editing the grants
@callback(
    Output("selected-grant-id", "data"),
    Input({'type': 'edit-btn', 'index': ALL}, 'n_clicks'),
    prevent_initial_call=True
)
def store_edit_id(n_clicks_list):
    if not any(n_clicks_list):
        raise PreventUpdate
    triggered_id = ctx.triggered_id
    return triggered_id["index"]

@callback(
    Output("redirect-location", "href"),
    Input("selected-grant-id", "data"),
    prevent_initial_call=True
)
def redirect_to_edit_page(grant_id):
    if grant_id:
        return f"/home/generate-grants?edit_id={grant_id}"
    raise PreventUpdate
